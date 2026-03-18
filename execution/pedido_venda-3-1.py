from flask import Flask, render_template, request, send_file, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os
import logging
from logging.handlers import RotatingFileHandler
from consulta_dolar_bacen import obter_fechamento_dolar
from consulta_receitaws import consultar_receitaws
from consulta_cnpja import consultar_cnpja
from tabela import processar_tabela_precos, escrever_tabela_no_excel  # Importa as funções

import os

base_dir = os.path.dirname(os.path.dirname(__file__))
app = Flask(__name__, template_folder=os.path.join(base_dir, 'templates'))

# Configuração de Logs para o Agente ler e auto-corrigir
log_dir = os.path.join(base_dir, '.tmp')
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, 'system.log')

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(name)s - %(message)s',
    handlers=[
        RotatingFileHandler(log_file, maxBytes=1048576, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def limpar_cnpj(cnpj):
    # Remove todos os caracteres não numéricos
    return ''.join(filter(str.isdigit, cnpj))

# Função para buscar os dados e escrever no Excel
def fetch_data_and_write_to_excel(cnpj_number, tipo, uso, data_pedido=None, prazo_pagamento=None, frete=None, st_difal=None, tabela_precos=None, cotacao_dolar=None, data_formatada=None):
    # Limpar o CNPJ para usar apenas números
    cnpj_number = limpar_cnpj(cnpj_number)
    
    # Consultar ReceitaWS
    dados_receitaws = consultar_receitaws(cnpj_number)
    if not dados_receitaws:
        return None

    # Consultar CNPJA passando o UF da ReceitaWS
    dados_cnpja = consultar_cnpja(cnpj_number, dados_receitaws['uf'])

    # Processar a data do pedido
    hoje = datetime.now()
    if data_pedido:
        try:
            data_pedido = datetime.strptime(data_pedido, '%Y-%m-%d')
        except ValueError:
            try:
                # also try dd/mm/yyyy just in case
                data_pedido = datetime.strptime(data_pedido, '%d/%m/%Y')
            except ValueError:
                data_pedido = hoje
    else:
        data_pedido = hoje


    # Abrir o arquivo Excel existente limpo dos templates
    base_dir = os.path.dirname(os.path.dirname(__file__))
    layout_path = os.path.join(base_dir, 'templates', 'layout_predefinido.xlsx')
    workbook = load_workbook(layout_path)
    worksheet = workbook.active

    # Preencher as células do Excel com os dados
    worksheet['C5'] = tipo
    worksheet['C6'] = uso
    worksheet['E8'] = dados_cnpja['number_ie']
    worksheet['E7'] = dados_receitaws['fantasia']
    worksheet['C7'] = dados_receitaws['nome']
    worksheet['C8'] = dados_receitaws['cnpj']
    worksheet['C9'] = dados_receitaws['logradouro_numero']
    worksheet['E9'] = dados_receitaws['complemento']
    worksheet['C10'] = dados_receitaws['municipio']
    worksheet['E10'] = dados_receitaws['uf']
    worksheet['C11'] = dados_receitaws['bairro']
    worksheet['E11'] = dados_receitaws['cep']
    worksheet['B47'] = cotacao_dolar
    worksheet['E6'] = data_pedido.strftime('%m/%d/%Y')
    
    # Preencher os novos campos
    worksheet['D21'] = prazo_pagamento
    worksheet['D22'] = frete
    worksheet['D23'] = st_difal

    # Escrever a tabela de preços no Excel
    escrever_tabela_no_excel(worksheet, tabela_precos, cotacao_dolar)

    # Salvar e fechar o arquivo Excel na pasta temporária
    tmp_path = os.path.join(base_dir, '.tmp', 'pedido_gerado.xlsx')
    os.makedirs(os.path.dirname(tmp_path), exist_ok=True)
    workbook.save(tmp_path)

    # Combinar todos os dados para retornar
    return {
        **dados_receitaws,
        **dados_cnpja,
        "data_formatada": data_formatada,
        "cotacao_venda": cotacao_dolar,
        "data_pedido": data_pedido.strftime('%d/%m/%Y'),
        "prazo_pagamento": prazo_pagamento,
        "frete": frete,
        "st_difal": st_difal,
        "tabela_precos": tabela_precos # Inclui a tabela de preços no retorno
    }

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        try:
            # Salvar a tabela mestra se foi enviada (overwrite)
            tabela_upload = request.files.get('tabela_mestra')
            if tabela_upload and tabela_upload.filename != '':
                tabela_mestra_path = os.path.join(base_dir, 'Tabela_de_Preços_Mestra.xlsx')
                tabela_upload.save(tabela_mestra_path)
                logger.info(f"Nova tabela mestra carregada: {tabela_upload.filename}")

            cnpj_number = request.form['cnpj']
            tipo = request.form['tipo']
            uso = request.form['uso']
            data_pedido = request.form.get('data_pedido')
            prazo_pagamento = request.form.get('prazo_pagamento')
            frete = request.form.get('frete')
            st_difal = request.form.get('st_difal')
            nome_aba_tabela = request.form.get('nome_aba_tabela')

            # Logica para salvar a ultima aba usada
            aba_cache_path = os.path.join(base_dir, '.tmp', 'ultima_aba.txt')
            if nome_aba_tabela and nome_aba_tabela.strip():
                with open(aba_cache_path, 'w', encoding='utf-8') as f:
                    f.write(nome_aba_tabela.strip())
            else:
                if os.path.exists(aba_cache_path):
                    with open(aba_cache_path, 'r', encoding='utf-8') as f:
                        nome_aba_tabela = f.read().strip()

            logger.info(f"Processando pedido para CNPJ {cnpj_number} na aba '{nome_aba_tabela}' e data {data_pedido}")

            # Consultar a cotação do dólar baseada na data do pedido
            data_formatada, cotacao_venda, hoje = obter_fechamento_dolar(data_pedido)

            # Processar a tabela de preços usando a função do arquivo separado, agora passando o nome da aba
            tabela_precos = processar_tabela_precos(request, cotacao_venda, nome_aba_tabela)

            data = fetch_data_and_write_to_excel(
                cnpj_number, 
                tipo, 
                uso, 
                request.form.get('data_pedido'), # we ensure we pass the string here
                prazo_pagamento,
                frete,
                st_difal,
                tabela_precos, # Passa a tabela de preços para a função
                cotacao_dolar=cotacao_venda,
                data_formatada=data_formatada
            )

            if data:
                logger.info("Pedido gerado com sucesso!")
                return render_template('result.html', data=data)
            else:
                logger.warning(f"Erro ao buscar os dados para o CNPJ {cnpj_number}.")
                return 'Erro ao buscar os dados. Verifique o CNPJ.'
        except Exception as e:
            logger.exception("Falha crítica durante o processamento do pedido.")
            return f"Ocorreu um erro interno: {str(e)}", 500
    else:
        return render_template('index.html')

@app.route('/download', methods=['POST'])
def download():
    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, '.tmp', 'pedido_gerado.xlsx')
    response = send_file(file_path, as_attachment=True, download_name='Pedido_de_venda_gerado.xlsx')
    return response

@app.route('/get_sheets', methods=['GET', 'POST'])
def get_sheets():
    base_dir = os.path.dirname(os.path.dirname(__file__))
    tabela_path = os.path.join(base_dir, 'Tabela_de_Preços_Mestra.xlsx')
    aba_cache_path = os.path.join(base_dir, '.tmp', 'ultima_aba.txt')
    
    if request.method == 'POST':
        tabela_upload = request.files.get('tabela_mestra')
        if tabela_upload and tabela_upload.filename != '':
            tabela_upload.save(tabela_path)
            # Ao carregar nova tabela, limpa o cache da última aba usada
            if os.path.exists(aba_cache_path):
                os.remove(aba_cache_path)
            
    try:
        workbook = load_workbook(tabela_path, read_only=True, data_only=True)
        sheets = workbook.sheetnames
        workbook.close()
        
        last_sheet = ""
        if os.path.exists(aba_cache_path):
            with open(aba_cache_path, 'r', encoding='utf-8') as f:
                last_sheet = f.read().strip()
                
        return jsonify({"sheets": sheets, "last_sheet": last_sheet})
    except Exception as e:
        return jsonify({"sheets": [], "error": str(e), "last_sheet": ""})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5008, debug=True)