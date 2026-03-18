import openpyxl
from openpyxl import load_workbook
import os

def buscar_preco_produto(codigo_produto, cotacao_dolar, nome_aba_tabela=None):
    """Busca o preço do produto na tabela de preços mestra."""
    try:
        # Each call to buscar_preco_produto re-opens and loads the workbook.
        # This is good for ensuring you get the latest data from the file on disk.
        base_dir = os.path.dirname(os.path.dirname(__file__))
        tabela_path = os.path.join(base_dir, 'Tabela_de_Preços_Mestra.xlsx')
        workbook = openpyxl.load_workbook(tabela_path, data_only=True)
        
        if nome_aba_tabela and nome_aba_tabela in workbook.sheetnames:
            sheets_to_search = [workbook[nome_aba_tabela]]
        else:
            sheets_to_search = workbook.worksheets

        # Convert the input codigo_produto to string for reliable comparison
        codigo_produto_str = str(codigo_produto).strip().lower()

        for worksheet in sheets_to_search:
            # Scan first few rows to find header
            header_row = None
            col_codigo, col_desc, col_p1, col_p2, col_p3 = None, None, None, None, None
            
            for r in range(1, min(worksheet.max_row, 10) + 1):
                # Helper to find column index by keywords
                def find_col(keywords):
                    for c in range(1, worksheet.max_column + 1):
                        val = str(worksheet.cell(row=r, column=c).value).strip().lower()
                        if val and any(kw in val for kw in keywords):
                            return c
                    return None
                
                cand_codigo = find_col(['código', 'codigo', 'pn', 'part number'])
                cand_desc = find_col(['descrição', 'descricao', 'produto'])
                cand_p1 = find_col(['1-5', '1 - 5', '1 a 5', '1 até 5'])
                cand_p2 = find_col(['6-10', '6 - 10', '6 a 10', '6 até 10'])
                cand_p3 = find_col(['10+', '11-', 'acima', '>10', '> 10'])
                
                if cand_codigo and (cand_desc or cand_p1):
                    header_row = r
                    col_codigo = cand_codigo
                    col_desc = cand_desc if cand_desc else (cand_codigo + 1)
                    col_p1 = cand_p1 if cand_p1 else (col_desc + 1)
                    col_p2 = cand_p2 if cand_p2 else (col_p1 + 1)
                    col_p3 = cand_p3 if cand_p3 else (col_p2 + 1)
                    break

            # Se não achar cabeçalho válido nesta aba, passa para a próxima aba
            if not col_codigo:
                continue

            for row in range(header_row + 1, worksheet.max_row + 1):
                codigo_sheet_cell = worksheet.cell(row=row, column=col_codigo).value
                if codigo_sheet_cell is None:
                    continue
                codigo_sheet_str = str(codigo_sheet_cell).strip().lower()

                if codigo_sheet_str == codigo_produto_str:
                    descricao = worksheet.cell(row=row, column=col_desc).value
                    preco1_cell = worksheet.cell(row=row, column=col_p1).value
                    preco2_cell = worksheet.cell(row=row, column=col_p2).value
                    preco3_cell = worksheet.cell(row=row, column=col_p3).value

                    def to_float_or_none(val):
                        if isinstance(val, (int, float)):
                            return float(val)
                        if isinstance(val, str):
                            try:
                                v = val.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                                import re
                                v = re.sub(r'[^\d,-]', '', val)
                                return float(v.replace(',', '.'))
                            except ValueError:
                                return None
                        return None

                    preco1_val = to_float_or_none(preco1_cell)
                    preco2_val = to_float_or_none(preco2_cell)
                    preco3_val = to_float_or_none(preco3_cell)

                    if preco1_val is None or preco2_val is None or preco3_val is None:
                        print(f"Aviso: Preço para o código {codigo_produto_str} contém valor inválido na aba '{worksheet.title}' linha {row}.")
                        return descricao, None, None, None

                    preco1 = round(cotacao_dolar * preco1_val, 2)
                    preco2 = round(cotacao_dolar * preco2_val, 2)
                    preco3 = round(cotacao_dolar * preco3_val, 2)

                    return descricao, preco1, preco2, preco3

        print(f"Debug: Produto com código '{codigo_produto_str}' não encontrado em nenhuma aba válida.")
        return None, None, None, None  # Produto não encontrado
    except FileNotFoundError:
        print(f"Arquivo '{tabela_path}' não encontrado.")
        return None, None, None, None
    except Exception as e:
        print(f"Erro inesperado em buscar_preco_produto para '{codigo_produto}': {e}")
        # import traceback
        # traceback.print_exc() # For more detailed error during debugging
        return None, None, None, None

def processar_tabela_precos(request, cotacao_venda, nome_aba_tabela=None):
    """Processa os dados da tabela de preços vindos do formulário."""
    tabela_precos = []
    print(f"Debug: Cotacao venda recebida em processar_tabela_precos: {cotacao_venda}")
    try:
        # Ensure cotacao_venda is a float for calculations
        cotacao_venda_float = float(cotacao_venda)
    except (ValueError, TypeError):
        print("Erro: cotacao_venda inválida em processar_tabela_precos.")
        # Handle case where cotacao_venda might not be convertible to float
        # For all items, prices will effectively be 0 or error state
        cotacao_venda_float = 0 # Or raise an error, or return empty table with errors

    for i in range(5): # Assuming you have 5 rows in your HTML form
        quantidade_str = request.form.get(f'quantidade_{i}')
        codigo_produto_form = request.form.get(f'codigo_produto_{i}')

        print(f"Debug: Linha {i}: Quantidade='{quantidade_str}', Codigo Produto='{codigo_produto_form}'")

        # Only process if both quantity and product code are provided and not empty
        if quantidade_str and codigo_produto_form and quantidade_str.strip() and codigo_produto_form.strip():
            try:
                quantidade = int(quantidade_str)
                if quantidade <= 0: # Skip if quantity is zero or negative
                    print(f"Debug: Linha {i}: Quantidade zero ou negativa, pulando.")
                    tabela_precos.append({
                        'quantidade': None, 'codigo_produto': None, 'descricao': None,
                        'preco_unitario': None, 'total': None
                    })
                    continue

                # Call buscar_preco_produto for each item
                descricao, preco1, preco2, preco3 = buscar_preco_produto(codigo_produto_form, cotacao_venda_float, nome_aba_tabela)
                print(f"Debug: Linha {i}: Resultado buscar_preco_produto: Desc='{descricao}', P1='{preco1}', P2='{preco2}', P3='{preco3}'")


                if preco1 is not None and preco2 is not None and preco3 is not None:
                    if quantidade < 6:
                        preco_unitario = preco1
                    elif 6 <= quantidade <= 10:
                        preco_unitario = preco2
                    else: # quantidade > 10
                        preco_unitario = preco3

                    total = quantidade * preco_unitario
                    tabela_precos.append({
                        'quantidade': quantidade,
                        'codigo_produto': codigo_produto_form,
                        'descricao': descricao,
                        'preco_unitario': preco_unitario,
                        'total': total
                    })
                else:
                    # Product code found but prices were invalid, or product not found
                    tabela_precos.append({
                        'quantidade': quantidade,
                        'codigo_produto': codigo_produto_form,
                        'descricao': descricao if descricao else 'Produto ou preço não encontrado/inválido',
                        'preco_unitario': 0, # Or None
                        'total': 0 # Or None
                    })
            except ValueError: # Catch error if quantidade_str is not a valid integer
                print(f"Debug: Linha {i}: Quantidade '{quantidade_str}' inválida.")
                tabela_precos.append({
                    'quantidade': None, # Or original string
                    'codigo_produto': codigo_produto_form,
                    'descricao': 'Erro: Quantidade inválida',
                    'preco_unitario': None,
                    'total': None
                })
        else:
            # Append a blank entry if not all data is present for this row
            tabela_precos.append({
                'quantidade': None, 'codigo_produto': None, 'descricao': None,
                'preco_unitario': None, 'total': None
            })
    return tabela_precos

def escrever_tabela_no_excel(worksheet, tabela_precos, cotacao_dolar): # Note: cotacao_dolar is not used here but was in your original signature
    """Escreve os dados da tabela de preços no arquivo Excel, incluindo as fórmulas."""
    if tabela_precos:
      linha_excel = 27  # Linha inicial no Excel
      for item in tabela_precos:
          if item['quantidade'] and item['codigo_produto']:
              quantidade = item['quantidade']
              codigo_produto = item['codigo_produto']
              descricao = item['descricao']
              preco_unitario = item['preco_unitario']

              worksheet[f'A{linha_excel}'] = quantidade
              worksheet[f'B{linha_excel}'] = codigo_produto
              worksheet[f'C{linha_excel}'] = descricao
              if preco_unitario is not None:
                worksheet[f'D{linha_excel}'] = preco_unitario
                worksheet[f'E{linha_excel}'].value = f"=A{linha_excel}*D{linha_excel}"
              else:
                worksheet[f'D{linha_excel}'] = 'Valor Indisponível'
                worksheet[f'E{linha_excel}'] = 'Valor Indisponível'

              linha_excel += 1  # Avança para a próxima linha